from openpyxl import load_workbook
from openpyxl import Workbook

class ExcelManager:
    def __init__(self, name = 0, size = 16):
        self.name = "code/result/name{}.xlsx".format(name)
        self.size = size
        self.open_book()

    def open_book(self):
        sheet_name = "S{}"
        try:
            self.wb = load_workbook(filename = self.name, data_only=True)
        except:
            self.wb = Workbook()
            self.wb.save(filename = self.name)
            self.wb = load_workbook(filename=self.name, data_only = True)
            for i in range(self.size):
                self.wb.create_sheet(sheet_name.format(i), i)
            self.wb.create_sheet("Sheet_father")
            self.wb.create_sheet("Sheet_mother")
            self.wb.save(self.name)

        self.letter = [self.wb[sheet_name.format(i)] for i in range(self.size)]
        self.letter_best = self.wb["Sheet"]
        self.letter_father = self.wb["Sheet_father"]
        self.letter_mother = self.wb["Sheet_mother"]

    def read(self, number = 0):
        i = 0
        last_row = []
        first = 0
        for row in self.letter[number].iter_rows(min_row = 1):
            i += 1
            last_row = []
            first = 0
            for j in range(len(row)):
                if j == 0:
                    first = row[j].value
                else:
                    last_row.append(row[j].value)
        return first, last_row

    def write_data2D(self, index = 0, hight=1, weights=[[0, 0], [1, 1]]):
        new_mas = []
        for i in range(len(weights)):
            for j in range(len(weights[i])):
                new_mas += weights[i][j]
        self.write_data(self.letter[index], hight, new_mas)

    def write_data2D_father(self, hight=1, weights=[[0, 0], [1, 1]]):
        new_mas = []
        for i in range(len(weights)):
            for j in range(len(weights[i])):
                new_mas += weights[i][j]
        self.write_data(self.letter_father, hight, new_mas)
    def write_data2D_mother(self, hight=1, weights=[[0, 0], [1, 1]]):
        new_mas = []
        for i in range(len(weights)):
            for j in range(len(weights[i])):
                new_mas += weights[i][j]
        self.write_data(self.letter_mother, hight, new_mas)
    def write_data2D_best(self, hight=1, weights=[[0, 0], [1, 1]]):
        new_mas = []
        for i in range(len(weights)):
            for j in range(len(weights[i])):
                new_mas += weights[i][j]
        self.write_data(self.letter_best, hight, new_mas)

    def write_data(self, letter, hight = 1, weights = [0]):
        data_to_end = [hight] + weights
        letter.append(data_to_end)
        self.save_data()
    def save_data(self):
        self.wb.save(self.name)
        self.open_book()